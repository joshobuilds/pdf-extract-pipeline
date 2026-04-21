"""Output writers. JSON for pipelines, Excel for humans."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pdf_extract.extractor import ExtractionResult
from pdf_extract.schema import Schema


def write_json(results: list[ExtractionResult], out_path: str | Path) -> Path:
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [
        {
            "source": r.source_path,
            "confidence": r.confidence,
            "missing_fields": r.missing_fields,
            "warnings": r.warnings,
            "data": r.data,
        }
        for r in results
    ]
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return path


def write_excel(
    results: list[ExtractionResult],
    schema: Schema,
    out_path: str | Path,
) -> Path:
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    data_sheet = wb.active
    data_sheet.title = "Extracted"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    columns = ["source", *schema.field_names(), "confidence", "missing_fields", "warnings"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = data_sheet.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, result in enumerate(results, start=2):
        data_sheet.cell(row=row_idx, column=1, value=Path(result.source_path).name)
        for col_idx, field_name in enumerate(schema.field_names(), start=2):
            value = result.data.get(field_name)
            data_sheet.cell(row=row_idx, column=col_idx, value=_format_cell(value))
        offset = len(schema.field_names()) + 2
        data_sheet.cell(row=row_idx, column=offset, value=round(result.confidence, 3))
        data_sheet.cell(row=row_idx, column=offset + 1, value="; ".join(result.missing_fields))
        data_sheet.cell(row=row_idx, column=offset + 2, value="; ".join(result.warnings))

    for col_idx in range(1, len(columns) + 1):
        data_sheet.column_dimensions[get_column_letter(col_idx)].width = 22
    data_sheet.freeze_panes = "A2"

    wb.save(path)
    return path


def _format_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, default=str)
    return value
